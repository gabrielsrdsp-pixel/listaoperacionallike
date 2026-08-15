import streamlit as st
import math
import pandas as pd
from datetime import datetime
import os
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Configuração da página e layout wide
st.set_page_config(
    page_title="Sistemas Integrados | CFTV & Alarmes",
    page_icon="🛡️",
    layout="wide"
)

# Estilização CSS Personalizada para o Layout Dark Profissional
st.markdown("""
    <style>
    .stApp {
        background-color: #0E1117;
        color: #E6EDF3;
    }
    .main-header {
        font-size: 2rem;
        font-weight: 800;
        color: #FFFFFF;
        margin-bottom: 0px;
        letter-spacing: 0.5px;
    }
    .sub-header {
        font-size: 0.95rem;
        color: #8B949E;
        margin-bottom: 10px;
    }
    .section-header {
        font-size: 1.3rem;
        font-weight: 700;
        color: #FFFFFF;
        margin-top: 10px;
        margin-bottom: 15px;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    [data-testid="stMetric"] {
        background-color: #0D1117;
        border: 1px solid #1F6FEB;
        border-radius: 8px;
        padding: 15px;
    }
    [data-testid="stMetricLabel"] {
        color: #8B949E !important;
        font-size: 0.85rem !important;
        font-weight: 600;
    }
    [data-testid="stMetricValue"] {
        color: #58A6FF !important;
        font-size: 1.8rem !important;
        font-weight: 700;
    }
    [data-testid="stVerticalBlock"] > div[data-testid="stBlock"] {
        border-radius: 8px;
    }
    .dica-box {
        background-color: #0D1117;
        border-left: 4px solid #1F6FEB;
        padding: 12px;
        border-radius: 4px;
        font-size: 0.85rem;
        color: #C9D1D9;
        margin-bottom: 15px;
    }
    .stDownloadButton button {
        width: 100%;
        background-color: #238636;
        color: white;
        font-weight: bold;
        border: none;
        padding: 12px;
        border-radius: 6px;
    }
    .stDownloadButton button:hover {
        background-color: #2EA043;
    }
    </style>
""", unsafe_allow_html=True)

if "historico_downloads" not in st.session_state:
    st.session_state["historico_downloads"] = []

# ==========================================
# FUNÇÃO PREMIUM DE IMPRESSÃO EXCEL (DINÂMICA DE COLUNAS)
# ==========================================
def gerar_excel_formatado(df_dados, modulo, tecnico, cliente, data_hora, protocolo):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Lista {modulo}"

    # Habilitar linhas de grade na impressão
    ws.views.sheetView[0].showGridLines = True

    # Definir Estilos
    fonte_titulo = Font(name="Calibri", size=18, bold=True, color="1F6FEB")
    fonte_sub = Font(name="Calibri", size=10, italic=True, color="555555")
    fonte_label = Font(name="Calibri", size=11, bold=True, color="1F2937")
    fonte_valor = Font(name="Calibri", size=11, color="1F2937")
    fonte_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fonte_corpo = Font(name="Calibri", size=11, color="000000")
    fonte_ass = Font(name="Calibri", size=10, bold=True, color="333333")

    fill_blue_header = PatternFill(start_color="1F6FEB", end_color="1F6FEB", fill_type="solid")
    fill_info_box = PatternFill(start_color="F0F6FF", end_color="F0F6FF", fill_type="solid")
    fill_row_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    border_fina = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    border_linha_ass = Border(top=Side(style='medium', color='1F2937'))

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

    # 1. TÍTULO E PROTOCOLO
    ws["A1"] = f"LISTA DE SEPARAÇÃO — {modulo.upper()}"
    ws["A1"].font = fonte_titulo
    ws["A1"].alignment = Alignment(vertical="center")

    ws["A2"] = f"Protocolo: {protocolo}"
    ws["A2"].font = fonte_sub

    # 2. BLOCO DE INFORMAÇÕES
    info_rows = [
        ("Responsável pela Separação:", tecnico if tecnico else "Não Informado"),
        ("Cliente / Empresa:", cliente if cliente else "Não Informado"),
        ("Data e Hora da Emissão:", data_hora)
    ]

    # Obter apenas as colunas ativas na tabela visual
    headers = list(df_dados.columns)
    total_cols = max(len(headers), 2)

    for idx, (label, val) in enumerate(info_rows, start=4):
        ws.row_dimensions[idx].height = 22
        ws[f"A{idx}"] = f"  {label}"
        ws[f"A{idx}"].font = fonte_label
        ws[f"A{idx}"].fill = fill_info_box
        
        ws[f"B{idx}"] = val
        ws[f"B{idx}"].font = fonte_valor
        ws[f"B{idx}"].fill = fill_info_box

        # Preencher fundo até a largura total disponível das colunas ativas
        for c_idx in range(1, total_cols + 1):
            c_letter = openpyxl.utils.get_column_letter(c_idx)
            ws[f"{c_letter}{idx}"].fill = fill_info_box
            ws[f"{c_letter}{idx}"].border = border_fina

    ws.row_dimensions[7].height = 15

    # 3. CABEÇALHO DINÂMICO DA TABELA (Apenas colunas presentes no DataFrame)
    ws.row_dimensions[8].height = 26
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=8, column=col_idx, value=str(header).upper())
        cell.font = fonte_header
        cell.fill = fill_blue_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_fina

    # 4. PREENCHIMENTO DOS DADOS
    row_idx = 9
    for i, (_, row) in enumerate(df_dados.iterrows()):
        ws.row_dimensions[row_idx].height = 24
        usar_zebra = (i % 2 == 1)

        for col_idx, col_name in enumerate(headers, start=1):
            val = row[col_name]
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if pd.notna(val) else "")
            cell.font = fonte_corpo
            cell.border = border_fina
            
            if usar_zebra:
                cell.fill = fill_row_zebra

            # Alinhamentos dinâmicos
            if "qtd" in col_name.lower() or "quantidade" in col_name.lower():
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        row_idx += 1

    # 5. BLOCO DE ASSINATURAS ADAPTÁVEL
    row_idx += 2
    ws.row_dimensions[row_idx].height = 10

    half_col = max(1, total_cols // 2)

    # Assinatura 1
    ws.cell(row=row_idx, column=1).border = border_linha_ass
    for c in range(1, half_col + 1):
        ws.cell(row=row_idx, column=c).border = border_linha_ass
    if half_col > 1:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=half_col)
        ws.merge_cells(start_row=row_idx+1, start_column=1, end_row=row_idx+1, end_column=half_col)
    
    cell_ass1 = ws.cell(row=row_idx+1, column=1, value="RESPONSÁVEL PELA SEPARAÇÃO")
    cell_ass1.font = fonte_ass
    cell_ass1.alignment = Alignment(horizontal="center", vertical="center")

    # Assinatura 2
    start_col2 = half_col + 1
    if start_col2 <= total_cols:
        for c in range(start_col2, total_cols + 1):
            ws.cell(row=row_idx, column=c).border = border_linha_ass
        if total_cols > start_col2:
            ws.merge_cells(start_row=row_idx, start_column=start_col2, end_row=row_idx, end_column=total_cols)
            ws.merge_cells(start_row=row_idx+1, start_column=start_col2, end_row=row_idx+1, end_column=total_cols)

        cell_ass2 = ws.cell(row=row_idx+1, column=start_col2, value="CONFERIDO / RECEBIDO POR")
        cell_ass2.font = fonte_ass
        cell_ass2.alignment = Alignment(horizontal="center", vertical="center")

    # Ajuste automático das larguras de colunas
    for col_idx, col_name in enumerate(headers, start=1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        max_len = max(len(str(val or '')) for val in df_dados[col_name])
        max_len = max(max_len, len(str(col_name)))
        ws.column_dimensions[col_letter].width = max(max_len + 5, 18)

    # IMPRESSÃO A4 PERFECT-FIT
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.6
    ws.page_margins.bottom = 0.6

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# ==========================================
# CABEÇALHO SUPERIOR (LOGO + TÍTULOS)
# ==========================================
col_logo, col_titulo = st.columns([1, 5])

with col_logo:
    caminho_pasta = os.path.dirname(os.path.abspath(__file__)) if '__file__' in globals() else os.getcwd()
    caminho_png = os.path.join(caminho_pasta, "logo.png")
    caminho_jpg = os.path.join(caminho_pasta, "logo.jpg")

    logo_carregada = False
    if os.path.exists(caminho_png):
        st.image(caminho_png, width=120)
        logo_carregada = True
    elif os.path.exists(caminho_jpg):
        st.image(caminho_jpg, width=120)
        logo_carregada = True

    if not logo_carregada:
        st.markdown("🛡️ **SISTEMAS INTEGRADOS**")

with col_titulo:
    st.markdown('<p class="main-header">LIKE SOLUTIONS</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">📍 Sistema Integrado de Dimensionamento — CFTV & Alarmes</p>', unsafe_allow_html=True)

# ==========================================
# NAVEGAÇÃO DE ABAS
# ==========================================
aba_cftv, aba_alarme = st.tabs(["Módulo CFTV (Câmeras)", "Módulo de Alarmes"])
data_hora_atual = datetime.now().strftime("%d/%m/%Y às %H:%M:%S")

# ==========================================
# ABA 1: CFTV (CÂMERAS)
# ==========================================
with aba_cftv:
    st.markdown('<p class="section-header">PARÂMETROS DO SISTEMA DE CFTV</p>', unsafe_allow_html=True)

    c_box1, c_box2 = st.columns(2)

    with c_box1:
        with st.container(border=True):
            st.markdown("**CÂMERAS & DISPOSIÇÃO**")
            qtd_cameras = st.number_input("Quantidade de Câmeras:", min_value=0, value=10, step=1, key="cftv_qtd")
            distancia_media = st.number_input("Distância média por câmera (m):", min_value=0.0, value=20.0, step=5.0, key="cftv_dist")

    with c_box2:
        with st.container(border=True):
            st.markdown("**INFRAESTRUTURA & TECNOLOGIA**")
            tipo_inst = st.selectbox("Tecnologia / Alimentação:", ["1 - PoE (Power over Ethernet)", "2 - Normal (Fonte 12V / BNC / Balun)"], key="cftv_tipo")
            tipo_fixacao = st.selectbox("Superfície de Fixação:", ["Alvenaria / Tijolo / Concreto", "Estrutura Metálica / Perfilado / Drywall"], key="cftv_fix")
            parafusos_por_camera = st.number_input("Parafusos por câmera:", min_value=1, max_value=10, value=2, step=1, key="cftv_paraf")

    if qtd_cameras > 0:
        parafusos = qtd_cameras * parafusos_por_camera
        buchas = qtd_cameras * parafusos_por_camera
        caixas_sobrepor = qtd_cameras
        metragem_cabo = math.ceil((qtd_cameras * distancia_media) * 1.1)

        if "Estrutura Metálica" in tipo_fixacao:
            item_fixador = "Parafuso Autobrocante (metal)"
            qtd_fixador = f"{parafusos} un"
            obs_fixador = "Fixação direta sem bucha"
        else:
            item_fixador = "Parafuso + Bucha S6/S8"
            qtd_fixador = f"{parafusos} un / {buchas} un"
            obs_fixador = f"{parafusos_por_camera} conjuntos por caixa"

        dados_cftv = [
            {"Categoria": "Equipamentos", "Item / Insumo": "Câmeras de Segurança", "Quantidade": f"{qtd_cameras} un", "Observação": "Pontos de monitoramento"},
            {"Categoria": "Acessórios", "Item / Insumo": "Caixas de Sobrepor (VBOX)", "Quantidade": f"{caixas_sobrepor} un", "Observação": "Proteção para conectores"},
            {"Categoria": "Fixação", "Item / Insumo": item_fixador, "Quantidade": qtd_fixador, "Observação": obs_fixador},
            {"Categoria": "Cabeamento", "Item / Insumo": "Cabo UTP Cat5e / Cat6", "Quantidade": f"~{metragem_cabo} m", "Observação": "Inclui 10% de margem de sobra"},
        ]

        if "1 - PoE" in tipo_inst:
            sw_desc = "Switch PoE 4 a 8 Portas" if qtd_cameras <= 4 else ("Switch PoE 8 a 10 Portas" if qtd_cameras <= 8 else ("Switch PoE 16 Portas" if qtd_cameras <= 16 else "Switch PoE 24 Portas"))
            conectores = (qtd_cameras * 2) + 4
            dados_cftv.extend([
                {"Categoria": "Conectores", "Item / Insumo": "Conectores RJ45 Macho", "Quantidade": f"{conectores} un", "Observação": "Inclui margem de erro"},
                {"Categoria": "Ativos de Rede", "Item / Insumo": sw_desc, "Quantidade": "1 un", "Observação": "Alimentação e dados no mesmo cabo"},
                {"Categoria": "Energia", "Item / Insumo": 'Régua de Tomadas Padrão 19"', "Quantidade": "1 un", "Observação": "Para Switch/NVR"}
            ])
        else:
            baluns = qtd_cameras * 2
            conectores_p4 = qtd_cameras * 2
            amperagem = math.ceil(qtd_cameras * 1.0)
            dados_cftv.extend([
                {"Categoria": "Conectores", "Item / Insumo": "Baluns de Vídeo HD", "Quantidade": f"{baluns} un ({qtd_cameras} pares)", "Observação": "Sinal de vídeo"},
                {"Categoria": "Conectores", "Item / Insumo": "Conectores P4 (Macho/Fêmea)", "Quantidade": f"{conectores_p4} un ({qtd_cameras} pares)", "Observação": "Alimentação das câmeras"},
                {"Categoria": "Energia", "Item / Insumo": f"Fonte Colméia 12V ({amperagem}A)", "Quantidade": "1 un", "Observação": "Dimensionada em 1A por câmera"},
                {"Categoria": "Conectores", "Item / Insumo": "Conectores RJ45", "Quantidade": "4 un", "Observação": "Para DVR e Roteador"}
            ])

        st.markdown("<br>", unsafe_allow_html=True)
        m1, m2, m3 = st.columns(3)
        m1.metric("Total de Câmeras:", f"{qtd_cameras} un")
        m2.metric("Tecnologia:", "PoE (IP)" if "1 - PoE" in tipo_inst else "Analógica (12V)")
        m3.metric("Cabo Estimado:", f"~{metragem_cabo} m")

        st.markdown("""
            <div class="dica-box">
                💡 <b>Dicas para Editar a Tabela:</b><br>
                • <b>Excluir item:</b> Selecione a caixinha à esquerda do item e pressione <code>Delete</code> (ou clique na lixeira).<br>
                • <b>Ocultar/Excluir Colunas:</b> Se você deletar ou ocultar colunas na tabela, elas também não aparecerão na impressão.
            </div>
        """, unsafe_allow_html=True)

        df_cftv_base = pd.DataFrame(dados_cftv)
        df_cftv_editado = st.data_editor(
            df_cftv_base,
            use_container_width=True,
            hide_index=True,
            num_rows="dynamic",
            key="editor_cftv_dynamic"
        )

        st.markdown("<br>", unsafe_allow_html=True)
        with st.container(border=True):
            st.markdown("**📋 IDENTIFICAÇÃO DA SEPARAÇÃO DE MATERIAL (OBRIGATÓRIO)**")
            col_cftv_tec, col_cftv_emp, col_cftv_dt = st.columns(3)
            with col_cftv_tec:
                tec_cftv = st.text_input("👤 Nome do Responsável / Separador: *", placeholder="Ex: João Silva", key="cftv_tec")
            with col_cftv_emp:
                emp_cftv = st.text_input("🏢 Nome da Empresa / Cliente: *", placeholder="Ex: Cliente Exemplo", key="cftv_emp")
            with col_cftv_dt:
                st.text_input("📅 Data e Hora:", value=data_hora_atual, disabled=True, key="cftv_dt")

        download_cftv_liberado = bool(tec_cftv.strip() and emp_cftv.strip())

        if not download_cftv_liberado:
            st.warning("⚠️ Preencha o **Nome do Responsável** e o **Nome da Empresa** para liberar o download.")

        protocolo_cftv = f"PROT-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        excel_cftv = gerar_excel_formatado(
            df_cftv_editado, "CFTV", tec_cftv, emp_cftv, data_hora_atual, protocolo_cftv
        )

        def registrar_download_cftv():
            st.session_state["historico_downloads"].append({
                "Protocolo": protocolo_cftv,
                "Data/Hora": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                "Módulo": "CFTV",
                "Empresa/Cliente": emp_cftv,
                "Responsável/Separador": tec_cftv
            })

        st.download_button(
            label="📥 BAIXAR TABELA DE CFTV EM EXCEL (FORMATO IMPRESSÃO PREMIUM)",
            data=excel_cftv,
            file_name=f"Lista_CFTV_{protocolo_cftv}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="down_cftv",
            disabled=not download_cftv_liberado,
            on_click=registrar_download_cftv
        )

# ==========================================
# ABA 2: ALARMES
# ==========================================
with aba_alarme:
    st.markdown('<p class="section-header">PARÂMETROS DO SISTEMA DE ALARME</p>', unsafe_allow_html=True)

    box1, box2 = st.columns(2)

    with box1:
        with st.container(border=True):
            st.markdown("**SENSORES**")
            qtd_ivp = st.number_input("Sensores de Presença (IVP):", min_value=0, value=10, step=1, key="al_ivp")
            qtd_mag = st.number_input("Sensores Magnéticos (Porta/Janela):", min_value=0, value=4, step=1, key="al_mag")
            dist_media_alarme = st.number_input("Distância média por sensor (m):", min_value=0.0, value=10.0, step=5.0, key="al_dist")
            qtd_sirenes = st.number_input("Quantidade de Sirenes:", min_value=1, value=1, step=1, key="al_sir")

    with box2:
        with st.container(border=True):
            st.markdown("**INFRAESTRUTURA & TIPO**")
            qtd_teclados = st.number_input("Quantidade de Teclados:", min_value=1, value=1, step=1, key="al_tec")
            tipo_fix_alarme = st.selectbox("Fixação dos Sensores:", ["Estrutura Metálica / MDF / Drywall", "Alvenaria / Tijolo / Concreto"], key="al_fix")
            com_fio = st.radio("Tipo dos Sensores:", ["Sensores Sem Fio", "Sensores Com Fio"], key="al_fio")
            parafusos_sensor = st.number_input("Parafusos por sensor:", min_value=1, max_value=6, value=2, step=1, key="al_paraf")

    total_sensores = qtd_ivp + qtd_mag

    if total_sensores > 0:
        tot_parafusos_al = (total_sensores + qtd_sirenes + qtd_teclados) * parafusos_sensor
        tot_buchas_al = tot_parafusos_al
        cabo_4vias = math.ceil((total_sensores * dist_media_alarme) * 1.1) if "Com Fio" in com_fio else 0

        if total_sensores <= 4:
            central_rec = "Central de Alarme 4 Zonas"
        elif total_sensores <= 8:
            central_rec = "Central de Alarme 8 Zonas (Expandível)"
        elif total_sensores <= 16:
            central_rec = "Central de Alarme 16 Zonas (Modular)"
        else:
            central_rec = "Central de Alarme de Grande Porte / Barramento"

        if "Estrutura Metálica" in tipo_fix_alarme:
            fix_desc = f"Parafuso Autobrocante ({tot_parafusos_al} un)"
            obs_fix_al = "Sem necessidade de buchas"
        else:
            fix_desc = f"Parafuso + Bucha S5/S6 ({tot_parafusos_al} parafusos / {tot_buchas_al} buchas)"
            obs_fix_al = f"{parafusos_sensor} por periférico"

        dados_alarme = [
            {"Categoria": "Central & Comando", "Item / Insumo": central_rec, "Quantidade": "1 un", "Observação": f"Atende aos {total_sensores} sensores"},
            {"Categoria": "Central & Comando", "Item / Insumo": "Teclado de Operação", "Quantidade": f"{qtd_teclados} un", "Observação": "Ativação e armação"},
            {"Categoria": "Sinalização", "Item / Insumo": "Sirene Piezoelétrica / Alta Potência", "Quantidade": f"{qtd_sirenes} un", "Observação": "Alerta sonoro"},
            {"Categoria": "Energia / Backup", "Item / Insumo": "Bateria Selada 12V 7Ah", "Quantidade": "1 un", "Observação": "Autonomia na falta de energia"},
            {"Categoria": "Fixação", "Item / Insumo": fix_desc, "Quantidade": f"{tot_parafusos_al} un", "Observação": obs_fix_al},
        ]

        if qtd_ivp > 0:
            dados_alarme.append({"Categoria": "Detecção", "Item / Insumo": "Sensor de Presença (IVP)", "Quantidade": f"{qtd_ivp} un", "Observação": "Detecção de movimento"})
        
        if qtd_mag > 0:
            dados_alarme.append({"Categoria": "Detecção", "Item / Insumo": "Sensor Magnético (Porta/Janela)", "Quantidade": f"{qtd_mag} un", "Observação": "Proteção perimetral"})

        if "Com Fio" in com_fio:
            dados_alarme.append({"Categoria": "Cabeamento", "Item / Insumo": "Cabo de Alarme 4 Vias CCI/Multicores", "Quantidade": f"~{cabo_4vias} m", "Observação": "Inclui 10% de sobra"})
            dados_alarme.append({"Categoria": "Acessórios", "Item / Insumo": "Articuladores / Suportes para IVP", "Quantidade": f"{qtd_ivp} un", "Observação": "Ajuste de ângulo dos sensores"})
        else:
            dados_alarme.append({"Categoria": "Baterias", "Item / Insumo": "Baterias CR2032 / Lithio (Para sensores)", "Quantidade": f"{total_sensores} un", "Observação": "Alimentação dos sensores sem fio"})

        st.markdown("<br>", unsafe_allow_html=True)
        a1, a2, a3 = st.columns(3)
        a1.metric("Total de Sensores:", f"{total_sensores} un")
        a2.metric("Tipo de Sistema:", "Sem Fio" if "Sem Fio" in com_fio else "Com Fio")
        a3.metric("Cabo Estimado:", f"~{cabo_4vias} m" if "Com Fio" in com_fio else "N/A (Sem Fio)")

        st.markdown("""
            <div class="dica-box">
                💡 <b>Dicas para Editar a Tabela:</b><br>
                • <b>Excluir item:</b> Selecione a caixinha à esquerda do item e pressione <code>Delete</code> (ou clique na lixeira).<br>
                • <b>Ocultar/Excluir Colunas:</b> Se você deletar ou ocultar colunas na tabela, elas também não aparecerão na impressão.
            </div>
        """, unsafe_allow_html=True)

        df_alarme_base = pd.DataFrame(dados_alarme)
        df_alarme_editado = st.data_editor(
            df_alarme_base,
            use_container_width=True,
            hide_index=True,
            num_rows="dynamic",
            key="editor_alarme_dynamic"
        )

        st.markdown("<br>", unsafe_allow_html=True)
        with st.container(border=True):
            st.markdown("**📋 IDENTIFICAÇÃO DA SEPARAÇÃO DE MATERIAL (OBRIGATÓRIO)**")
            col_al_tec, col_al_emp, col_al_dt = st.columns(3)
            with col_al_tec:
                tec_alarme = st.text_input("👤 Nome do Responsável / Separador: *", placeholder="Ex: João Silva", key="al_tec_input")
            with col_al_emp:
                emp_alarme = st.text_input("🏢 Nome da Empresa / Cliente: *", placeholder="Ex: Cliente Exemplo", key="al_emp_input")
            with col_al_dt:
                st.text_input("📅 Data e Hora:", value=data_hora_atual, disabled=True, key="al_dt_input")

        download_alarme_liberado = bool(tec_alarme.strip() and emp_alarme.strip())

        if not download_alarme_liberado:
            st.warning("⚠️ Preencha o **Nome do Responsável** e o **Nome da Empresa** para liberar o download.")

        protocolo_alarme = f"PROT-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        excel_alarme = gerar_excel_formatado(
            df_alarme_editado, "Alarme", tec_alarme, emp_alarme, data_hora_atual, protocolo_alarme
        )

        def registrar_download_alarme():
            st.session_state["historico_downloads"].append({
                "Protocolo": protocolo_alarme,
                "Data/Hora": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                "Módulo": "Alarme",
                "Empresa/Cliente": emp_alarme,
                "Responsável/Separador": tec_alarme
            })

        st.download_button(
            label="📥 BAIXAR TABELA DE ALARMES EM EXCEL (FORMATO IMPRESSÃO PREMIUM)",
            data=excel_alarme,
            file_name=f"Lista_Alarme_{protocolo_alarme}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="down_alarme",
            disabled=not download_alarme_liberado,
            on_click=registrar_download_alarme
        )

# ==========================================
# PAINEL DE HISTÓRICO DE DOWNLOADS
# ==========================================
st.divider()
st.markdown('<p class="section-header">📜 HISTÓRICO DE LISTAS BAIXADAS (PROTOCOLOS)</p>', unsafe_allow_html=True)

if len(st.session_state["historico_downloads"]) == 0:
    st.caption("Nenhuma lista foi baixada nesta sessão ainda.")
else:
    df_hist = pd.DataFrame(st.session_state["historico_downloads"])
    st.dataframe(df_hist, use_container_width=True, hide_index=True)
    
    if st.button("🗑️ Limpar Histórico de Downloads", key="btn_clear_hist"):
        st.session_state["historico_downloads"] = []
        st.rerun()